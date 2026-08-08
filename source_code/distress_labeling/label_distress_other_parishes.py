import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SRC_DIR = r"H:\stats america\distress_checking"
OUT_DIR = r"H:\stats america\distress_checking\distress data labeled"
EXCLUDE_FILES = {"East_Baton_Rouge_LA_Tract_distress_download.csv"}

UNEMPLOYMENT_DIFF_COL = "Threshold Calculation"
INCOME_SHARE_COL = "Threshold Calculation2"

LIGHT_RED = "FFC7CE"
DARK_RED = "FF0000"
LIGHT_GREEN = "C6EFCE"

label_columns = [
    "Distressed by Unemployment Rate",
    "Distressed by Per Capita Income Share",
    "Distressed by Both Unemployment Rate and Per Capita Income Share",
    "Non-Distressed",
]

light_red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
dark_red_fill = PatternFill(start_color=DARK_RED, end_color=DARK_RED, fill_type="solid")
light_green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

os.makedirs(OUT_DIR, exist_ok=True)

for src_path in glob.glob(os.path.join(SRC_DIR, "*.csv")):
    if os.path.basename(src_path) in EXCLUDE_FILES:
        continue

    base_name = os.path.splitext(os.path.basename(src_path))[0]
    xlsx_out = os.path.join(OUT_DIR, f"{base_name}_Labeled.xlsx")

    df = pd.read_csv(src_path)

    unemployment_diff = pd.to_numeric(df[UNEMPLOYMENT_DIFF_COL], errors="coerce")
    income_share = pd.to_numeric(df[INCOME_SHARE_COL], errors="coerce")

    distressed_unemployment = unemployment_diff >= 1.0
    distressed_income = income_share <= 80.0

    df["Distressed by Unemployment Rate"] = distressed_unemployment.map({True: "Yes", False: "No"})
    df["Distressed by Per Capita Income Share"] = distressed_income.map({True: "Yes", False: "No"})
    df["Distressed by Both Unemployment Rate and Per Capita Income Share"] = (distressed_unemployment & distressed_income).map({True: "Yes", False: "No"})
    df["Non-Distressed"] = (~distressed_unemployment & ~distressed_income).map({True: "Yes", False: "No"})

    df.to_excel(xlsx_out, index=False)

    # Apply cell fill colors to the new label columns in the xlsx copy
    wb = load_workbook(xlsx_out)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_index = {name: header.index(name) + 1 for name in label_columns}

    for row in range(2, ws.max_row + 1):
        for name, col in col_index.items():
            cell = ws.cell(row=row, column=col)
            if cell.value == "Yes":
                if name == "Distressed by Both Unemployment Rate and Per Capita Income Share":
                    cell.fill = dark_red_fill
                elif name == "Non-Distressed":
                    cell.fill = light_green_fill
                else:
                    cell.fill = light_red_fill

    wb.save(xlsx_out)

    print("Saved XLSX:", xlsx_out)
